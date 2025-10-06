import pandas as pd
import requests, logging, io
import redis.asyncio as redis
from typing import List, Dict
from openpyxl import load_workbook
from fastapi import FastAPI, HTTPException, Request
from datetime import datetime, timezone, timedelta
from configs import LISTINGS_URL, REDIS_URL, fetch_apify_jobs, hash_job, sendEmailAlert, timestamp_to_datetime




logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)




# DNA
app = FastAPI(
    title='Grad INTERN List v1',
    description='Grad INTERN List v1',
    version='0.0.1',
)




# On-Start
@app.on_event("startup")
async def startup_event():
    app.state.redis = redis.from_url(f"{REDIS_URL}")
    logger.info("Connected to Redis")




# On-Destroy
@app.on_event("shutdown")
async def shutdown_event():
    await app.state.redis.close()
    logger.info("Disconnected from Redis")




# health
@app.get('/health')
async def check_alive(request: Request):
    return {'message': 'alive'}




@app.get("/recent_jobs", response_model=Dict)
async def get_recent_jobs(minutes: int = 60):
    # Fetch Simplify listings
    try:
        response = requests.get(LISTINGS_URL, timeout=20)
        response.raise_for_status()
        listings = response.json()
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Failed to fetch listings.json: {e}")

    cutoff = datetime.now(timezone.utc) - timedelta(minutes=minutes)

    recent_jobs = []
    for job in listings:
        if not job.get("active", False) or not job.get("is_visible", True):
            continue
        date_posted_ts = job.get("date_posted")
        if not date_posted_ts:
            continue
        try:
            date_posted = timestamp_to_datetime(int(date_posted_ts))
        except Exception:
            continue
        if date_posted >= cutoff:
            recent_jobs.append(job)


    # Fetch LinkedIn jobs
    apify_jobs = fetch_apify_jobs()
    apify_normalized = []
    new_jobs = []
    redis = app.state.redis

    for job in apify_jobs:
        normalized = {
            "Title": job.get("title"),
            "Company Name": job.get("companyName"),
            "Location": job.get("location"),
            "Posted time": job.get("postedTime"),
            "Published at": job.get("publishedAt"),
            "Job Url": job.get("jobUrl"),
            "Applications count": job.get("applicationsCount"),
            "Employment type": job.get("contractType"),
        }
        apify_normalized.append(normalized)


        # --- Check Redis to send only new jobs ---
        job_hash = hash_job(normalized)
        job_url = normalized.get("Job Url", "")
        
        # Check if this job hash already exists in Redis
        exists = await redis.exists(job_hash)
        logger.info(f"Job hash exists in Redis: {exists}")
        
        if not exists:
            new_jobs.append(normalized)
            # Store job hash as key and job URL as value with 3-day expiration
            await redis.setex(job_hash, 3 * 24 * 60 * 60, job_url)
            logger.info(f"Added new job: {normalized.get('Title', 'No Title')} - Total new jobs: {len(new_jobs)}")
        else:
            logger.info(f"Job already exists, skipping: {normalized.get('Title', 'No Title')}")


    logger.info(f"New jobs found: {len(new_jobs)}")

    # Save Excel with clickable URLs
    if recent_jobs or new_jobs:
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            if recent_jobs:
                pd.DataFrame(recent_jobs).to_excel(writer, sheet_name="recent_jobs", index=False)
            if new_jobs:
                pd.DataFrame(new_jobs).to_excel(writer, sheet_name="new_jobs", index=False)
        excel_buffer.seek(0)

        # Make Job Url clickable in both sheets
        wb = load_workbook(excel_buffer)
        for sheet_name in ["recent_jobs", "new_jobs"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for idx, cell in enumerate(ws[1], start=1):
                    if cell.value == "Job Url":
                        job_url_col = idx
                        break
                else:
                    job_url_col = None
                if job_url_col:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=job_url_col)
                        if cell.value:
                            cell.hyperlink = cell.value
                            cell.style = "Hyperlink"
        # Save workbook back to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        if len(recent_jobs) != 0 or len(new_jobs) != 0:
            try:
                sendEmailAlert(
                    recent_jobs,
                    attachment=excel_buffer,
                    attachment_name="linkedin_jobs_24h.xlsx"
                )
                logger.info("Email sent!")
            except Exception as e:
                logger.error(f"Failed to send email: {e}")

    return {
        "recent_jobs_count": len(recent_jobs),
        "total_apify_jobs": len(apify_normalized),
        "new_apify_jobs": len(new_jobs)
    }